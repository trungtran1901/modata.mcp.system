"""
mcp_servers/attendance_analytics_server.py  (v3 — full logic update)

Các thay đổi so với v2:
  1. Lấy nhân viên theo phong_cap_1 / phong_cap_2 / phong_cap_3 (phân cấp đúng)
  2. Giờ làm việc theo dia_diem_lam_viec:
       - "Hòa Lạc" / "Hoa Lac" → 08:30 ~ 16:30
       - Các địa điểm khác       → 08:30 ~ 17:30
  3. Nghỉ giữa giờ 12:00 ~ 13:15 (75 phút) — trừ khỏi thực tế làm việc
  4. Nghỉ phép nửa ngày (0.5 công) và xuyên ngày tính đúng số công
  5. Đơn đi muộn/về sớm: tối đa 2 đơn/kỳ, tổng phút ≤ 60
     (vượt quá → không tính miễn phút, tính trừ bình thường)
  6. Đơn "Cập nhật công" đã duyệt → ghi đè giờ vào/ra ngày đó
  7. Sheet chi tiết ngày cập nhật đủ: loại nghỉ, phút muộn, note

CẤU HÌNH THỜI GIAN (dễ điều chỉnh):
  Tìm block "── TIME CONFIG ──" bên dưới để sửa mốc giờ.

Tools (prefix att_ana_):
  get_attendance_data          — trả JSON đầy đủ
  export_attendance_excel      — xuất Excel 2 sheet
  compute_and_export           — all-in-one
  send_attendance_report       — tính + xuất + gửi mail
"""
from __future__ import annotations

import json
import logging
import os
import re
from datetime import datetime, timedelta, date, time as dtime
from typing import Any, Optional

from fastmcp import FastMCP

from app.db.mongo import get_db
from utils.session import get_session_context

logger = logging.getLogger(__name__)
mcp = FastMCP("modata-attendance-analytics")


# ═══════════════════════════════════════════════════════════════
# ── TIME CONFIG ── (điều chỉnh mốc giờ tại đây)
# ═══════════════════════════════════════════════════════════════

# Ca hành chính thông thường (tất cả địa điểm NGOẠI TRỪ Hòa Lạc)
NORMAL_WORK_START = dtime(8, 30)   # 08:30
NORMAL_WORK_END   = dtime(17, 30)  # 17:30

# Ca Hòa Lạc
HAALAC_WORK_START = dtime(8, 30)   # 08:30
HAALAC_WORK_END   = dtime(16, 30)  # 16:30

# Địa điểm Hòa Lạc (so sánh lowercase, không dấu)
_HOA_LAC_KEYWORDS = {"hoa lac", "hoalac", "hòa lạc", "hoa lạc"}

# Nghỉ giữa giờ
BREAK_START = dtime(12, 0)    # 12:00
BREAK_END   = dtime(13, 15)   # 13:15
BREAK_MINUTES = 75            # = (13:15 - 12:00) tính bằng phút

# Số ngày công chuẩn / kỳ
STD_CONG = 26

# Đơn đi muộn/về sớm được miễn: tối đa 2 đơn/kỳ VÀ tổng phút ≤ 60
LATE_EXEMPT_MAX_ORDERS  = 2
LATE_EXEMPT_MAX_MINUTES = 60

# ═══════════════════════════════════════════════════════════════

WEEKDAY_VI   = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]
WEEKDAY_FULL = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def _session_ok(session_id: str) -> bool:
    return bool(get_session_context(session_id).accessible_instance_names())


def _ev(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, dict):
        return v.get("label") or v.get("value")
    if isinstance(v, list):
        items = [_ev(i) for i in v if i is not None]
        items = [i for i in items if i is not None]
        return items[0] if len(items) == 1 else (items or None)
    if type(v).__name__ in ("ObjectId", "datetime"):
        return str(v)
    return v


def _utc_to_vn(dt: Any) -> datetime | None:
    """MongoDB UTC datetime → Vietnam time (UTC+7)."""
    if dt is None:
        return None
    if hasattr(dt, "strftime"):
        return dt + timedelta(hours=7)
    return None


def _hm(dt: datetime | None) -> tuple[int, int] | None:
    return (dt.hour, dt.minute) if dt else None


def _fmt_time(hm: tuple | None) -> str:
    if hm is None:
        return "-:-"
    return f"{hm[0]:02d}:{hm[1]:02d}"


def _is_hoa_lac(location_str: str | None) -> bool:
    """Kiểm tra địa điểm có phải Hòa Lạc không."""
    if not location_str:
        return False
    return location_str.lower().strip() in _HOA_LAC_KEYWORDS


def _work_hours_for_location(location: str | None) -> tuple[dtime, dtime]:
    """Trả về (work_start, work_end) theo địa điểm."""
    if _is_hoa_lac(location):
        return HAALAC_WORK_START, HAALAC_WORK_END
    return NORMAL_WORK_START, NORMAL_WORK_END


def _work_minutes_for_location(location: str | None) -> int:
    """Tổng phút làm việc chuẩn (trừ nghỉ giữa giờ)."""
    start, end = _work_hours_for_location(location)
    total = (end.hour * 60 + end.minute) - (start.hour * 60 + start.minute)
    # Trừ nghỉ giữa giờ nếu nằm trong ca
    if start <= BREAK_START and BREAK_END <= end:
        total -= BREAK_MINUTES
    return total


def _actual_work_minutes(fi: tuple | None, lo: tuple | None,
                          work_start: dtime, work_end: dtime) -> int:
    """
    Tính phút làm việc thực tế của 1 ngày,
    đã trừ nghỉ giữa giờ và clamp vào ca.
    """
    if fi is None and lo is None:
        return 0

    # Clamp vào ca
    clamp_start_h, clamp_start_m = work_start.hour, work_start.minute
    clamp_end_h,   clamp_end_m   = work_end.hour,   work_end.minute

    fi_min = fi[0] * 60 + fi[1] if fi else clamp_start_h * 60 + clamp_start_m
    lo_min = lo[0] * 60 + lo[1] if lo else clamp_end_h   * 60 + clamp_end_m

    # Không vượt ngoài ca
    fi_min = max(fi_min, clamp_start_h * 60 + clamp_start_m)
    lo_min = min(lo_min, clamp_end_h   * 60 + clamp_end_m)

    if lo_min <= fi_min:
        return 0

    # Trừ nghỉ giữa giờ nếu overlap
    brk_s = BREAK_START.hour * 60 + BREAK_START.minute
    brk_e = BREAK_END.hour   * 60 + BREAK_END.minute
    overlap = max(0, min(lo_min, brk_e) - max(fi_min, brk_s))

    return lo_min - fi_min - overlap


def _minutes_late(fi: tuple | None, work_start: dtime) -> int:
    """Phút đi muộn so với work_start."""
    if fi is None:
        return 0
    ws_min = work_start.hour * 60 + work_start.minute
    return max(0, fi[0] * 60 + fi[1] - ws_min)


def _minutes_early(lo: tuple | None, work_end: dtime) -> int:
    """Phút về sớm so với work_end."""
    if lo is None:
        return 0
    we_min = work_end.hour * 60 + work_end.minute
    return max(0, we_min - (lo[0] * 60 + lo[1]))


# ─────────────────────────────────────────────────────────────
# NGHỈ PHÉP NỬA NGÀY — tính số công
# ─────────────────────────────────────────────────────────────

def _calc_leave_cong(tu_vn: datetime, den_vn: datetime,
                     work_start: dtime, work_end: dtime) -> float:
    """
    Tính số ngày công nghỉ phép của 1 đơn trong 1 ngày cụ thể.

    Trả về: 1.0, 0.5, hoặc 0.0
    Logic:
      - Thời điểm nghỉ trong nửa đầu ca (buổi sáng) hoặc nửa sau (chiều)
        → 0.5 công
      - Trùm cả ca → 1.0 công
      Mốc nửa ngày = trung điểm giữa work_start và BREAK_START
                    (thực tế dùng 12:00 làm ranh giới sáng/chiều)
    """
    ws_min = work_start.hour * 60 + work_start.minute
    we_min = work_end.hour   * 60 + work_end.minute
    noon   = BREAK_END.hour  * 60 + BREAK_END.minute  # 13:15 — cuối nghỉ trưa

    # Thời gian nghỉ quy về phút trong ngày (UTC+7 đã xử lý)
    tu_min  = tu_vn.hour  * 60 + tu_vn.minute
    den_min = den_vn.hour * 60 + den_vn.minute

    # Overlap với ca làm việc
    eff_start = max(tu_min,  ws_min)
    eff_end   = min(den_min, we_min)

    if eff_end <= eff_start:
        return 0.0

    # Nếu đơn chỉ nằm trong buổi sáng (kết thúc <= 12:00)
    if den_min <= BREAK_START.hour * 60:
        return 0.5

    # Nếu đơn chỉ nằm trong buổi chiều (bắt đầu >= 13:15)
    if tu_min >= noon:
        return 0.5

    # Nếu trùm cả 2 buổi
    return 1.0


def _calc_leave_days_for_period(
    tu_vn: datetime, den_vn: datetime,
    target_date: date,
    so_ngay_nghi: float,
    work_start: dtime, work_end: dtime,
) -> float:
    """
    Tính số công nghỉ phép cho ngày target_date từ 1 đơn nghỉ.
    Xử lý đơn xuyên ngày và nửa ngày.
    """
    # Nếu so_ngay_nghi = 0.5 → nửa ngày
    if so_ngay_nghi == 0.5:
        return 0.5

    tu_date  = tu_vn.date()
    den_date = den_vn.date()

    # Đơn 1 ngày: kiểm tra giờ cụ thể
    if tu_date == den_date == target_date:
        return _calc_leave_cong(tu_vn, den_vn, work_start, work_end)

    # Đơn xuyên ngày: ngày đầu (từ giờ nộp đến hết ngày) và ngày cuối (từ đầu ngày đến giờ kết thúc)
    if target_date == tu_date:
        # Buổi chiều hôm đó (từ tu_vn đến hết ca)
        we_fake = datetime(target_date.year, target_date.month, target_date.day,
                           work_end.hour, work_end.minute)
        cong = _calc_leave_cong(tu_vn, we_fake, work_start, work_end)
        return cong

    if target_date == den_date:
        # Buổi sáng hôm sau (từ đầu ca đến den_vn)
        ws_fake = datetime(target_date.year, target_date.month, target_date.day,
                           work_start.hour, work_start.minute)
        cong = _calc_leave_cong(ws_fake, den_vn, work_start, work_end)
        return cong

    # Ngày giữa của đơn xuyên nhiều ngày → 1 công
    return 1.0


# ─────────────────────────────────────────────────────────────
# DATA LOADERS
# ─────────────────────────────────────────────────────────────

def _get_period_dates(year: int, month: int) -> tuple[date, date]:
    """Kỳ chấm công: 26/M-1 → 25/M."""
    if month == 1:
        start = date(year - 1, 12, 26)
    else:
        start = date(year, month - 1, 26)
    end = date(year, month, 25)
    return start, end


def _get_all_days_in_period(period_start: date, period_end: date) -> list[date]:
    result = []
    cur = period_start
    while cur <= period_end:
        result.append(cur)
        cur += timedelta(days=1)
    return result


def _get_work_days_in_period(period_start: date, period_end: date,
                              off_weekdays: set[int], holidays: set[str]) -> list[date]:
    result = []
    cur = period_start
    while cur <= period_end:
        ds = cur.strftime("%Y-%m-%d")
        if cur.weekday() not in off_weekdays and ds not in holidays:
            result.append(cur)
        cur += timedelta(days=1)
    return result


def _get_holidays_in_period(period_start: date, period_end: date,
                              company_code: str) -> set[str]:
    db = get_db()
    from_utc = datetime(period_start.year, period_start.month, period_start.day) - timedelta(hours=7)
    to_utc   = datetime(period_end.year,   period_end.month,   period_end.day,   23, 59, 59)
    docs = list(db["instance_data_ngay_nghi_le"].find({
        "is_deleted":   {"$ne": True},
        "company_code": company_code,
        "tu_ngay":  {"$lte": to_utc},
        "den_ngay": {"$gte": from_utc},
    }))
    result = set()
    for doc in docs:
        tu_vn  = _utc_to_vn(doc.get("tu_ngay"))
        den_vn = _utc_to_vn(doc.get("den_ngay"))
        if tu_vn and den_vn:
            cur = tu_vn.date()
            end = den_vn.date()
            while cur <= end:
                if period_start <= cur <= period_end:
                    result.add(cur.strftime("%Y-%m-%d"))
                cur += timedelta(days=1)
    return result


def _get_holiday_names(period_start: date, period_end: date,
                        company_code: str) -> dict[str, str]:
    db = get_db()
    from_utc = datetime(period_start.year, period_start.month, period_start.day) - timedelta(hours=7)
    to_utc   = datetime(period_end.year,   period_end.month,   period_end.day,   23, 59, 59)
    docs = list(db["instance_data_ngay_nghi_le"].find({
        "is_deleted":   {"$ne": True},
        "company_code": company_code,
        "tu_ngay":  {"$lte": to_utc},
        "den_ngay": {"$gte": from_utc},
    }))
    result: dict[str, str] = {}
    for doc in docs:
        ten    = doc.get("ten_ngay_nghi", "Nghỉ lễ")
        tu_vn  = _utc_to_vn(doc.get("tu_ngay"))
        den_vn = _utc_to_vn(doc.get("den_ngay"))
        if tu_vn and den_vn:
            cur = tu_vn.date()
            end = den_vn.date()
            while cur <= end:
                result[cur.strftime("%Y-%m-%d")] = ten
                cur += timedelta(days=1)
    return result


def _get_off_weekdays(company_code: str) -> set[int]:
    db = get_db()
    docs = list(db["instance_data_ngay_nghi_tuan"].find({
        "is_deleted":   {"$ne": True},
        "is_active":    {"$ne": False},
        "company_code": company_code,
    }))
    _MAP = {
        "thứ 2": 0, "thứ hai": 0, "t2": 0,
        "thứ 3": 1, "thứ ba": 1,  "t3": 1,
        "thứ 4": 2, "thứ tư": 2,  "t4": 2,
        "thứ 5": 3, "thứ năm": 3, "t5": 3,
        "thứ 6": 4, "thứ sáu": 4, "t6": 4,
        "thứ 7": 5, "thứ bảy": 5, "t7": 5,
        "chủ nhật": 6, "cn": 6,
    }
    off = set()
    for doc in docs:
        name = str(_ev(doc.get("loai_nghi_tuan")) or "").lower().strip()
        if name in _MAP:
            off.add(_MAP[name])
    return off


def _get_approved_leaves_in_period(
    username: str, period_start: date, period_end: date, company_code: str
) -> dict[str, list[dict]]:
    """
    Lấy đơn từ đã duyệt overlap với kỳ chấm công.
    Bao gồm: Nghỉ phép, Nghỉ ốm, Làm việc từ xa, Đề nghị đi công tác,
             Đi muộn về sớm, Cập nhật công
    Returns: {YYYY-MM-DD: [{"loai_don":..., "cong":float, "tu_vn":..., "den_vn":...,
                             "di_muon":int, "ve_som":int, "so_ngay_nghi":float,
                             "gio_vao":tuple|None, "gio_ra":tuple|None}]}
    """
    db = get_db()
    from_utc = datetime(period_start.year, period_start.month, period_start.day) - timedelta(hours=7)
    to_utc   = datetime(period_end.year,   period_end.month,   period_end.day,   23, 59, 59)

    # Query đơn từ: dùng cả tu_ngay và den_ngay để bắt đơn xuyên ngày + đơn 1 ngày
    docs = list(db["instance_data_danh_sach_quan_ly_don_xin_nghi"].find({
        "is_deleted":                 {"$ne": True},
        "company_code":               company_code,
        "nguoi_nop_don.value":        username,
        "trang_thai_phe_duyet.value": "Đã duyệt",
        "$or": [
            # Đơn có tu_ngay và den_ngay
            {"tu_ngay": {"$lte": to_utc}, "den_ngay": {"$gte": from_utc}},
            # Đơn Cập nhật công chỉ có den_ngay
            {"tu_ngay": None, "den_ngay": {"$gte": from_utc, "$lte": to_utc}},
        ],
    }))

    result: dict[str, list[dict]] = {}
    for doc in docs:
        loai         = doc.get("loai_don", "")
        so_ngay_nghi = float(doc.get("so_ngay_nghi") or 0)
        di_muon      = int(doc.get("di_muon_dau_ca") or 0)
        ve_som       = int(doc.get("ve_som_cuoi_ca") or 0)

        # Parse gio_vao / gio_ra từ đơn Cập nhật công
        gio_vao_raw = doc.get("gio_vao_dau_ca")
        gio_ra_raw  = doc.get("gio_ra_cuoi_ca")
        gio_vao_vn  = _utc_to_vn(gio_vao_raw)
        gio_ra_vn   = _utc_to_vn(gio_ra_raw)

        tu_vn  = _utc_to_vn(doc.get("tu_ngay"))
        den_vn = _utc_to_vn(doc.get("den_ngay"))

        # Đơn "Cập nhật công": chỉ có den_ngay (ngày cập nhật công)
        if loai == "Cập nhật công":
            if den_vn is None:
                continue
            ds = den_vn.date().strftime("%Y-%m-%d")
            if period_start <= den_vn.date() <= period_end:
                result.setdefault(ds, [])
                result[ds].append({
                    "loai_don":    loai,
                    "cong":        so_ngay_nghi,
                    "tu_vn":       tu_vn,
                    "den_vn":      den_vn,
                    "di_muon":     di_muon,
                    "ve_som":      ve_som,
                    "so_ngay_nghi": so_ngay_nghi,
                    "gio_vao":     _hm(gio_vao_vn),
                    "gio_ra":      _hm(gio_ra_vn),
                })
            continue

        if tu_vn is None and den_vn is None:
            continue
        if tu_vn is None:
            tu_vn = den_vn
        if den_vn is None:
            den_vn = tu_vn

        tu_date  = tu_vn.date()
        den_date = den_vn.date()

        cur = tu_date
        while cur <= den_date:
            if period_start <= cur <= period_end:
                ds = cur.strftime("%Y-%m-%d")
                result.setdefault(ds, [])
                result[ds].append({
                    "loai_don":     loai,
                    "cong":         0.0,   # sẽ tính sau theo giờ cụ thể
                    "tu_vn":        tu_vn,
                    "den_vn":       den_vn,
                    "di_muon":      di_muon,
                    "ve_som":       ve_som,
                    "so_ngay_nghi": so_ngay_nghi,
                    "gio_vao":      None,
                    "gio_ra":       None,
                })
            cur += timedelta(days=1)

    return result


def _get_late_orders_in_period(
    username: str, period_start: date, period_end: date, company_code: str
) -> list[dict]:
    """
    Lấy tất cả đơn Đi muộn/Về sớm đã duyệt trong kỳ.
    Dùng để áp dụng quy tắc: tối đa 2 đơn/kỳ, tổng ≤ 60 phút.
    """
    db = get_db()
    from_utc = datetime(period_start.year, period_start.month, period_start.day) - timedelta(hours=7)
    to_utc   = datetime(period_end.year,   period_end.month,   period_end.day,   23, 59, 59)

    docs = list(db["instance_data_danh_sach_quan_ly_don_xin_nghi"].find({
        "is_deleted":                 {"$ne": True},
        "company_code":               company_code,
        "nguoi_nop_don.value":        username,
        "loai_don":                   "Đi muộn, về sớm",
        "trang_thai_phe_duyet.value": "Đã duyệt",
        "$or": [
            {"tu_ngay": {"$lte": to_utc}, "den_ngay": {"$gte": from_utc}},
            {"tu_ngay": None, "den_ngay": {"$gte": from_utc, "$lte": to_utc}},
        ],
    }))
    return [
        {
            "di_muon": int(d.get("di_muon_dau_ca") or 0),
            "ve_som":  int(d.get("ve_som_cuoi_ca") or 0),
            "tu_vn":   _utc_to_vn(d.get("tu_ngay")),
            "den_vn":  _utc_to_vn(d.get("den_ngay")),
        }
        for d in docs
    ]


def _get_attendance_records(username: str, period_start: date,
                              period_end: date) -> dict[str, dict]:
    """Chấm công thô trong kỳ."""
    db = get_db()
    from_str = period_start.strftime("%Y-%m-%d")
    to_str   = period_end.strftime("%Y-%m-%d")

    ky_end_month = period_end.month
    ky_end_year  = period_end.year
    if ky_end_month == 1:
        ky_chot = f"{ky_end_year - 1}-12"
    else:
        ky_chot = f"{ky_end_year}-{ky_end_month - 1:02d}"

    docs = list(db["instance_data_lich_su_cham_cong_tong_hop_cong"].find({
        "is_deleted":    {"$ne": True},
        "ten_dang_nhap": username,
        "$or": [
            {"ngay_chot_cong": ky_chot},
            {"day": {"$gte": from_str, "$lte": to_str}},
        ],
    }))

    result: dict[str, dict] = {}
    for doc in docs:
        day_str = doc.get("day", "")
        if not day_str:
            continue
        try:
            d = date.fromisoformat(day_str)
            if not (period_start <= d <= period_end):
                continue
        except ValueError:
            continue
        fi = _hm(_utc_to_vn(doc.get("firstIn")))
        lo = _hm(_utc_to_vn(doc.get("lastOut")))
        result[day_str] = {"firstIn": fi, "lastOut": lo}
    return result


# ─────────────────────────────────────────────────────────────
# LẤY NHÂN VIÊN THEO PHÒNG BAN (phong_cap_1/2/3)
# ─────────────────────────────────────────────────────────────

def _get_employees(filter_type: str, filter_value: str,
                   company_code: str) -> list[dict]:
    """
    Lấy danh sách nhân viên theo bộ lọc.
    filter_type: "all" | "username" | "don_vi" | "phong_cap_1" | "phong_cap_2" | "phong_cap_3"

    Với "don_vi": tìm theo mã/tên phòng ban ở cả 3 cấp.
    Mỗi nhân viên trả về kèm thông tin phân cấp phòng ban và địa điểm làm việc.
    """
    db  = get_db()
    flt: dict = {
        "is_deleted":   {"$ne": True},
        "company_code": company_code,
        "trang_thai_lao_dong.value": "Đang làm việc",
    }

    if filter_type == "username":
        flt["ten_dang_nhap"] = filter_value

    elif filter_type in ("don_vi", "phong_cap_1", "phong_cap_2", "phong_cap_3"):
        # Tìm theo giá trị phòng ban ở các cấp
        or_conditions = []

        if filter_type in ("don_vi", "phong_cap_1"):
            or_conditions += [
                {"phong_cap_1.value": filter_value},
                {"phong_cap_1.label": {"$regex": filter_value, "$options": "i"}},
            ]
        if filter_type in ("don_vi", "phong_cap_2"):
            or_conditions += [
                {"phong_cap_2.value": filter_value},
                {"phong_cap_2.label": {"$regex": filter_value, "$options": "i"}},
            ]
        if filter_type in ("don_vi", "phong_cap_3"):
            or_conditions += [
                {"phong_cap_3.value": filter_value},
                {"phong_cap_3.label": {"$regex": filter_value, "$options": "i"}},
            ]

        if filter_type == "don_vi":
            # Fallback: tìm theo các field cũ nếu không khớp cấp
            or_conditions += [
                {"don_vi_cong_tac.value":       filter_value},
                {"don_vi_cong_tac.option.code": filter_value},
                {"phong_ban_phu_trach.value":   filter_value},
                {"path_don_vi_cong_tac": {"$regex": filter_value, "$options": "i"}},
            ]

        if or_conditions:
            flt["$or"] = or_conditions

    docs = list(db["instance_data_thong_tin_nhan_vien"].find(flt, {
        "ten_dang_nhap":    1,
        "ma_nhan_vien":     1,
        "ho_va_ten_co_dau": 1,
        "ho_va_ten":        1,
        "don_vi_cong_tac":  1,
        "phong_cap_1":      1,
        "phong_cap_2":      1,
        "phong_cap_3":      1,
        "vi_tri_cong_viec": 1,
        "dia_diem_lam_viec": 1,
    }).sort("ma_nhan_vien", 1))

    result = []
    for d in docs:
        # Phân cấp phòng ban
        p1 = d.get("phong_cap_1")
        p2 = d.get("phong_cap_2")
        p3 = d.get("phong_cap_3")

        phong_cap_1 = (_ev(p1) or "") if p1 else ""
        phong_cap_2 = (_ev(p2) or "") if p2 else ""
        phong_cap_3 = (_ev(p3) or "") if p3 else ""

        # Tên đơn vị hiển thị: cấp thấp nhất có giá trị
        don_vi_display = (
            phong_cap_3 or phong_cap_2 or phong_cap_1
            or _ev(d.get("don_vi_cong_tac")) or ""
        )

        # Địa điểm làm việc
        dia_diem = _ev(d.get("dia_diem_lam_viec")) or ""

        result.append({
            "username":    d.get("ten_dang_nhap", ""),
            "ma_nv":       d.get("ma_nhan_vien", ""),
            "ho_va_ten":   d.get("ho_va_ten_co_dau") or d.get("ho_va_ten", ""),
            "don_vi":      don_vi_display,
            "phong_cap_1": phong_cap_1,
            "phong_cap_2": phong_cap_2,
            "phong_cap_3": phong_cap_3,
            "vi_tri":      _ev(d.get("vi_tri_cong_viec")) or "",
            "dia_diem":    dia_diem,
        })
    return result


# ─────────────────────────────────────────────────────────────
# CORE CALCULATOR
# ─────────────────────────────────────────────────────────────

def _calc_one_employee(
    emp: dict,
    period_start: date, period_end: date,
    off_weekdays: set[int], holidays: set[str], holiday_names: dict[str, str],
    company_code: str,
) -> dict:
    """
    Tính toán đầy đủ cho 1 nhân viên.
    Dùng dia_diem từ hồ sơ nhân viên để xác định ca làm việc.
    """
    username = emp["username"]
    dia_diem = emp.get("dia_diem", "")
    work_start, work_end = _work_hours_for_location(dia_diem)
    std_work_min = _work_minutes_for_location(dia_diem)

    # --- Load data ---
    leaves     = _get_approved_leaves_in_period(username, period_start, period_end, company_code)
    att_recs   = _get_attendance_records(username, period_start, period_end)
    all_days   = _get_all_days_in_period(period_start, period_end)
    work_days  = _get_work_days_in_period(period_start, period_end, off_weekdays, holidays)
    work_days_set = {d.strftime("%Y-%m-%d") for d in work_days}

    # --- Xử lý đơn đi muộn/về sớm: quy tắc 2 đơn / 60 phút ---
    late_orders = _get_late_orders_in_period(username, period_start, period_end, company_code)
    late_exempt_map: dict[str, int] = {}  # {YYYY-MM-DD: phút được miễn}
    if len(late_orders) <= LATE_EXEMPT_MAX_ORDERS:
        total_late_phut = sum(o["di_muon"] + o["ve_som"] for o in late_orders)
        if total_late_phut <= LATE_EXEMPT_MAX_MINUTES:
            # Tất cả đơn được miễn → ghi vào map theo ngày
            for o in late_orders:
                ref_vn = o["tu_vn"] or o["den_vn"]
                if ref_vn:
                    ds = ref_vn.date().strftime("%Y-%m-%d")
                    late_exempt_map[ds] = late_exempt_map.get(ds, 0) + o["di_muon"] + o["ve_som"]

    # --- Phân loại ngày nghỉ ---
    nghi_phep_days: dict[str, float] = {}   # {ds: số_công_nghỉ}
    nghi_le_days:   set[str] = {d for d in holidays if d in work_days_set}
    wfh_days:       set[str] = set()
    cong_tac_days:  set[str] = set()
    cap_nhat_cong:  dict[str, dict] = {}    # {ds: {gio_vao, gio_ra}}

    for ds, leave_list in leaves.items():
        target_dt = date.fromisoformat(ds)
        for lv in leave_list:
            loai = lv["loai_don"]
            if loai in ("Nghỉ phép", "Nghỉ ốm"):
                tu_vn  = lv["tu_vn"]
                den_vn = lv["den_vn"]
                so_ngay = lv["so_ngay_nghi"]
                cong = _calc_leave_days_for_period(
                    tu_vn, den_vn, target_dt, so_ngay, work_start, work_end
                ) if tu_vn and den_vn else (so_ngay or 1.0)
                nghi_phep_days[ds] = nghi_phep_days.get(ds, 0.0) + cong

            elif loai == "Làm việc từ xa":
                wfh_days.add(ds)
            elif loai == "Đề nghị đi công tác":
                cong_tac_days.add(ds)
            elif loai == "Cập nhật công":
                cap_nhat_cong[ds] = {
                    "gio_vao": lv.get("gio_vao"),
                    "gio_ra":  lv.get("gio_ra"),
                }

    # --- Tính từng ngày ---
    tong_cong_thuc_te = 0.0
    nghi_kl_days: list[str] = []
    late_gt240_days:    list[str] = []
    late_60_240_days:   list[str] = []
    late_lt60_pool_min = 0

    daily_detail: list[dict] = []

    for d in all_days:
        ds       = d.strftime("%Y-%m-%d")
        weekday  = WEEKDAY_VI[d.weekday()]

        is_weekend = d.weekday() in off_weekdays
        is_holiday = ds in holidays
        is_leave   = ds in nghi_phep_days
        is_wfh     = ds in wfh_days
        is_ct      = ds in cong_tac_days
        has_cap_nhat = ds in cap_nhat_cong

        rec = att_recs.get(ds, {})
        fi  = rec.get("firstIn")
        lo  = rec.get("lastOut")

        # Ghi đè bởi đơn Cập nhật công
        if has_cap_nhat:
            cn = cap_nhat_cong[ds]
            if cn.get("gio_vao"):
                fi = cn["gio_vao"]
            if cn.get("gio_ra"):
                lo = cn["gio_ra"]

        note     = ""
        day_type = ""
        leave_cong = nghi_phep_days.get(ds, 0.0)

        if is_weekend:
            day_type = "Nghỉ tuần"

        elif is_holiday:
            day_type = f"Nghỉ lễ ({holiday_names.get(ds, 'Nghỉ lễ')})"

        elif is_leave and leave_cong >= 1.0:
            # Nghỉ phép cả ngày
            day_type = "Nghỉ phép/ốm"

        elif is_wfh:
            day_type = "WFH"
            tong_cong_thuc_te += 1.0

        elif is_ct:
            day_type = "Công tác"
            tong_cong_thuc_te += 1.0

        else:
            # Ngày làm việc — kiểm tra chấm công
            if fi is None and lo is None:
                if is_leave and leave_cong == 0.5:
                    # Nghỉ phép nửa ngày, không có CC → ghi nhận 0.5 công nghỉ
                    day_type = "Nghỉ phép nửa ngày"
                    tong_cong_thuc_te += 0.5
                else:
                    day_type = "Nghỉ không lương"
                    nghi_kl_days.append(ds)
            else:
                actual_min = _actual_work_minutes(fi, lo, work_start, work_end)

                if actual_min < 120:
                    day_type = "Nghỉ không lương"
                    nghi_kl_days.append(ds)
                else:
                    tong_cong_thuc_te += 1.0
                    if is_leave and leave_cong == 0.5:
                        day_type = "Nghỉ phép nửa ngày + Đi làm"
                    else:
                        day_type = "Đi làm"
                        if has_cap_nhat:
                            day_type = "Đi làm (cập nhật)"

                    # --- Phân tích đi muộn/về sớm ---
                    late_min  = _minutes_late(fi,  work_start)
                    early_min = _minutes_early(lo, work_end)

                    # Trừ đi phút được miễn theo đơn đã duyệt
                    exempt_min = late_exempt_map.get(ds, 0)
                    total_dev  = max(0, (late_min + early_min) - exempt_min)

                    if total_dev > 0:
                        if total_dev >= 240:
                            late_gt240_days.append(ds)
                            note = f"ĐM/VS {total_dev}ph (>4h, -1)"
                        elif total_dev >= 60:
                            late_60_240_days.append(ds)
                            note = f"ĐM/VS {total_dev}ph (1-4h, -0.5)"
                        else:
                            late_lt60_pool_min += total_dev
                            note = f"ĐM/VS {total_dev}ph"

                    if exempt_min > 0:
                        note = f"Miễn {exempt_min}ph đơn" + (f" | {note}" if note else "")

        # --- Note bổ sung nếu Cập nhật công ---
        if has_cap_nhat and day_type not in ("Nghỉ tuần", f"Nghỉ lễ ({holiday_names.get(ds, '')})"):
            note = "Cập nhật CC" + (f" | {note}" if note else "")

        daily_detail.append({
            "date":      ds,
            "day_no":    d.day,
            "weekday":   weekday,
            "check_in":  _fmt_time(fi),
            "check_out": _fmt_time(lo),
            "day_type":  day_type,
            "leave_cong": leave_cong,
            "note":      note,
        })

    # --- Tổng hợp ---
    nghi_phep_count = sum(
        min(v, 1.0) for ds, v in nghi_phep_days.items() if ds in work_days_set
    )
    # Nửa ngày: tính riêng
    nghi_phep_half  = sum(
        0.5 for ds, v in nghi_phep_days.items()
        if ds in work_days_set and v == 0.5
    )
    nghi_phep_full  = nghi_phep_count - nghi_phep_half

    nghi_le_count   = len([d for d in nghi_le_days   if d in work_days_set])
    wfh_count       = len([d for d in wfh_days       if d in work_days_set])
    cong_tac_count  = len([d for d in cong_tac_days  if d in work_days_set])
    nghi_kl_count   = len(nghi_kl_days)

    tong_cong_huong_luong = (
        STD_CONG
        - nghi_kl_count
        + nghi_phep_count   # bao gồm cả nửa ngày (0.5)
        + nghi_le_count
        + wfh_count
        + cong_tac_count
    )

    tru_sm = (
        len(late_gt240_days)  * 1.0
        + len(late_60_240_days) * 0.5
        + round(late_lt60_pool_min / 480, 4)
    )
    cong_tinh_luong = round(tong_cong_huong_luong - tru_sm, 4)

    return {
        "summary": {
            "so_cong_chuan":         STD_CONG,
            "nghi_phep":             round(nghi_phep_count, 2),
            "nghi_phep_nguyen_ngay": int(nghi_phep_full),
            "nghi_phep_nua_ngay":    round(nghi_phep_half, 2),
            "nghi_le":               nghi_le_count,
            "wfh":                   wfh_count,
            "cong_tac":              cong_tac_count,
            "nghi_khong_luong":      nghi_kl_count,
            "tong_cong_thuc_te":     round(tong_cong_thuc_te, 2),
            "tong_cong_huong_luong": round(tong_cong_huong_luong, 2),
            "dm_gt_4h":              len(late_gt240_days),
            "dm_1h_4h":              len(late_60_240_days),
            "phut_muon_lt_1h":       late_lt60_pool_min,
            "tru_sm":                round(tru_sm, 4),
            "cong_tinh_luong":       cong_tinh_luong,
            "dia_diem_lam_viec":     dia_diem,
            "gio_bat_dau_ca":        f"{work_start.hour:02d}:{work_start.minute:02d}",
            "gio_ket_thuc_ca":       f"{work_end.hour:02d}:{work_end.minute:02d}",
        },
        "daily_detail": daily_detail,
    }


def _recalculate_dependent_formulas(summary: dict, overrides_keys: set[str]) -> None:
    LATE_FIELDS  = {"dm_gt_4h", "dm_1h_4h", "phut_muon_lt_1h"}
    LEAVE_FIELDS = {"nghi_phep", "nghi_le", "wfh", "cong_tac", "nghi_khong_luong", "so_cong_chuan"}

    if LATE_FIELDS & overrides_keys and "tru_sm" not in overrides_keys:
        dm_gt  = float(summary.get("dm_gt_4h", 0))
        dm_mid = float(summary.get("dm_1h_4h", 0))
        phut   = float(summary.get("phut_muon_lt_1h", 0))
        summary["tru_sm"] = round(dm_gt * 1.0 + dm_mid * 0.5 + phut / 480, 4)

    if LEAVE_FIELDS & overrides_keys:
        sc   = float(summary.get("so_cong_chuan", STD_CONG))
        nkl  = float(summary.get("nghi_khong_luong", 0))
        np_  = float(summary.get("nghi_phep", 0))
        nl   = float(summary.get("nghi_le", 0))
        wfh  = float(summary.get("wfh", 0))
        ct   = float(summary.get("cong_tac", 0))
        summary["tong_cong_huong_luong"] = sc - nkl + np_ + nl + wfh + ct

    cong_huong = float(summary.get("tong_cong_huong_luong", STD_CONG))
    tru_sm     = float(summary.get("tru_sm", 0))
    summary["cong_tinh_luong"] = round(cong_huong - tru_sm, 4)


# ─────────────────────────────────────────────────────────────
# TOOLS
# ─────────────────────────────────────────────────────────────

@mcp.tool()
def get_attendance_data(
    session_id:   str,
    year_month:   str,
    filter_type:  str = "all",
    filter_value: str = "",
    company_code: str = "HITC",
) -> str:
    """
    Tính toán và trả về dữ liệu bảng chấm công đầy đủ dạng JSON.

    filter_type: "all" | "username" | "don_vi" | "phong_cap_1" | "phong_cap_2" | "phong_cap_3"
    filter_value: giá trị tương ứng (mã hoặc tên phòng ban, hoặc username)

    Giờ làm việc tự động theo dia_diem_lam_viec của từng nhân viên:
      - Hòa Lạc: 08:30 ~ 16:30
      - Còn lại:  08:30 ~ 17:30
    Nghỉ giữa giờ: 12:00 ~ 13:15 (75 phút).
    """
    if not _session_ok(session_id):
        return json.dumps({"error": "Session không hợp lệ."}, ensure_ascii=False)

    try:
        dt = datetime.strptime(year_month, "%Y-%m")
    except ValueError:
        return json.dumps({"error": "year_month phải có định dạng YYYY-MM"}, ensure_ascii=False)

    year, month = dt.year, dt.month
    period_start, period_end = _get_period_dates(year, month)

    holidays      = _get_holidays_in_period(period_start, period_end, company_code)
    holiday_names = _get_holiday_names(period_start, period_end, company_code)
    off_weekdays  = _get_off_weekdays(company_code)
    employees     = _get_employees(filter_type, filter_value, company_code)

    if not employees:
        return json.dumps({
            "error": f"Không tìm thấy nhân viên (filter_type={filter_type}, filter_value={filter_value})",
        }, ensure_ascii=False)

    results = []
    for emp in employees:
        try:
            calc = _calc_one_employee(
                emp, period_start, period_end,
                off_weekdays, holidays, holiday_names, company_code,
            )
            results.append({
                "username":    emp["username"],
                "ma_nv":       emp["ma_nv"],
                "ho_va_ten":   emp["ho_va_ten"],
                "don_vi":      emp["don_vi"],
                "phong_cap_1": emp["phong_cap_1"],
                "phong_cap_2": emp["phong_cap_2"],
                "phong_cap_3": emp["phong_cap_3"],
                "vi_tri":      emp["vi_tri"],
                "dia_diem":    emp["dia_diem"],
                "summary":     calc["summary"],
                "daily_detail": calc["daily_detail"],
            })
        except Exception as e:
            logger.warning("Calc error for %s: %s", emp["username"], e, exc_info=True)
            results.append({**emp, "error": str(e), "summary": {}, "daily_detail": []})

    ky_str = f"26/{period_start.month:02d}/{period_start.year} - 25/{period_end.month:02d}/{period_end.year}"

    return json.dumps({
        "year_month":  year_month,
        "period_info": {
            "ky_cham_cong":  ky_str,
            "period_start":  str(period_start),
            "period_end":    str(period_end),
            "total_days":    (period_end - period_start).days + 1,
            "holidays":      sorted(holidays),
            "holiday_names": holiday_names,
            "off_weekdays":  sorted(off_weekdays),
            "break_time":    f"{BREAK_START.strftime('%H:%M')} ~ {BREAK_END.strftime('%H:%M')}",
        },
        "so_nhan_vien": len(results),
        "employees":    results,
    }, ensure_ascii=False, default=str)


@mcp.tool()
def export_attendance_excel(
    session_id:              str,
    year_month:              str,
    filter_type:             str = "all",
    filter_value:            str = "",
    company_code:            str = "HITC",
    output_path:             str = "",
    extra_columns:           str = "",
    custom_formula_notes:    str = "",
    data_overrides:          str = "",
) -> str:
    """
    Xuất file Excel bảng chấm công tổng hợp.

    Sheet "Tổng hợp": các cột tổng hợp + công tính lương.
      Bao gồm cột "Nghỉ phép nửa ngày" riêng.
    Sheet "Chi tiết ngày": mỗi ngày = 1 cột (check-in / check-out + note).
    Sheet "Chú giải": màu sắc và công thức.

    data_overrides: JSON '{"mã_nv": {"tên_field": giá_trị}}'
      Công thức phụ thuộc tự động tính lại sau override.
    """
    if not _session_ok(session_id):
        return json.dumps({"error": "Session không hợp lệ."}, ensure_ascii=False)

    raw = json.loads(get_attendance_data(
        session_id=session_id, year_month=year_month,
        filter_type=filter_type, filter_value=filter_value,
        company_code=company_code,
    ))
    if "error" in raw:
        return json.dumps(raw, ensure_ascii=False)

    employees   = raw["employees"]
    period_info = raw["period_info"]
    ky_str      = period_info["ky_cham_cong"]
    period_start = date.fromisoformat(period_info["period_start"])
    period_end   = date.fromisoformat(period_info["period_end"])
    holidays     = set(period_info["holidays"])
    off_wdays    = set(period_info["off_weekdays"])

    # Áp dụng data_overrides
    if data_overrides:
        try:
            overrides = json.loads(data_overrides)
            for emp in employees:
                emp_override = overrides.get(emp.get("ma_nv", ""))
                if emp_override and isinstance(emp_override, dict):
                    summary = emp.setdefault("summary", {})
                    for field, val in emp_override.items():
                        try:
                            summary[field] = float(val) if isinstance(val, str) and val.replace(".", "", 1).lstrip("-").isdigit() else val
                        except Exception:
                            summary[field] = val
                    _recalculate_dependent_formulas(summary, set(emp_override.keys()))
        except Exception as e:
            logger.error("data_overrides parse error: %s", e, exc_info=True)

    # Parse extra columns
    extra_cols: dict[str, Any] = {}
    if extra_columns:
        try:
            parsed = json.loads(extra_columns)
            extra_cols = {c: "" for c in parsed} if isinstance(parsed, list) else parsed
        except Exception:
            pass

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return json.dumps({"error": "Thiếu openpyxl. Cài: pip install openpyxl"}, ensure_ascii=False)

    wb = Workbook()

    # ── Styles ─────────────────────────────────────────────
    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", start_color=hex_color)

    def _font(bold=False, color="000000", size=9) -> Font:
        return Font(bold=bold, color=color, name="Arial", size=size)

    def _border() -> Border:
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    C_HDR1 = "1F4E79"; C_HDR2 = "2E75B6"
    C_WE   = "D6E4F0"; C_HOL  = "FCE4D6"
    C_NP   = "E2EFDA"; C_NP_H = "C6EFC0"   # Nghỉ phép nửa ngày - xanh đậm hơn
    C_WFH  = "FFF2CC"; C_CT   = "EAD1DC"
    C_NKL  = "F4CCCC"; C_DL   = "DEEBF7"
    C_TOTAL = "FFFDE7"
    bd = _border()
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    ra = Alignment(horizontal="right",  vertical="center")

    # ═══════════════════════════════════════════
    # SHEET 1: TỔNG HỢP
    # ═══════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Tổng hợp"

    fixed_headers = [
        ("STT",                 5),
        ("Mã NV",               9),
        ("Họ và tên",           22),
        ("Phòng cấp 1",         18),
        ("Phòng cấp 2",         18),
        ("Phòng cấp 3",         18),
        ("Vị trí",              18),
        ("Địa điểm",            12),
        ("Công chuẩn",          9),
        ("Nghỉ phép\n(ngày)",   9),
        ("Nghỉ phép\n(nửa ngày)", 9),
        ("Nghỉ lễ",             7),
        ("WFH",                 7),
        ("Công tác",            8),
        ("Nghỉ KL",             8),
        ("Công TT",             8),
        ("Công HLương",         10),
        ("ĐM >4h\n(ngày)",      9),
        ("ĐM 1-4h\n(ngày)",     9),
        ("Phút muộn\n<1h",      9),
        ("Trừ SM",              8),
        ("Công tính lương",     13),
    ]
    all_headers = fixed_headers + [(k, 12) for k in extra_cols]
    total_cols = len(all_headers)
    end_col = get_column_letter(total_cols)

    ws1.merge_cells(f"A1:{end_col}1")
    ws1["A1"] = f"BẢNG CHẤM CÔNG TỔNG HỢP — {company_code}"
    ws1["A1"].font = _font(bold=True, color="1F4E79", size=14)
    ws1["A1"].alignment = ca

    ws1.merge_cells(f"A2:{end_col}2")
    ws1["A2"] = f"Kỳ chấm công: {ky_str}  |  Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1["A2"].font = _font(size=10)
    ws1["A2"].alignment = ca

    hdr_row = 4
    if custom_formula_notes:
        ws1.merge_cells(f"A3:{end_col}3")
        ws1["A3"] = f"Ghi chú: {custom_formula_notes}"
        ws1["A3"].font = _font(size=9, color="FF0000")
        ws1["A3"].alignment = la
        hdr_row = 5

    for ci, (lbl, wid) in enumerate(all_headers, 1):
        cell = ws1.cell(row=hdr_row, column=ci, value=lbl)
        cell.font      = _font(bold=True, color="FFFFFF", size=9)
        cell.fill      = _fill(C_HDR1)
        cell.alignment = ca
        cell.border    = bd
        ws1.column_dimensions[get_column_letter(ci)].width = wid
    ws1.row_dimensions[hdr_row].height = 45

    FIELD_MAP = [
        None,
        "ma_nv", "ho_va_ten", "phong_cap_1", "phong_cap_2", "phong_cap_3", "vi_tri", "dia_diem",
        "so_cong_chuan",
        "nghi_phep_nguyen_ngay",
        "nghi_phep_nua_ngay",
        "nghi_le", "wfh", "cong_tac",
        "nghi_khong_luong",
        "tong_cong_thuc_te",
        "tong_cong_huong_luong",
        "dm_gt_4h", "dm_1h_4h", "phut_muon_lt_1h",
        "tru_sm",
        "cong_tinh_luong",
    ]

    totals: dict[str, float] = {k: 0.0 for k in [
        "nghi_phep_nguyen_ngay", "nghi_phep_nua_ngay",
        "nghi_le", "wfh", "cong_tac", "nghi_khong_luong",
        "tong_cong_thuc_te", "tong_cong_huong_luong",
        "dm_gt_4h", "dm_1h_4h", "phut_muon_lt_1h", "tru_sm", "cong_tinh_luong",
    ]}

    TEXT_COLS = {1, 2, 3, 4, 5, 6, 7, 8}   # STT, mã, tên, phòng×3, vị trí, địa điểm
    NUM_INT_COLS  = {9, 10, 11, 12, 13, 14, 15, 18, 19, 20}
    NUM_DEC_COLS  = {16, 17, 21, 22}

    for idx, emp in enumerate(employees, 1):
        r = hdr_row + idx
        s = emp.get("summary", {})
        row_fill = _fill("EBF3FB") if idx % 2 == 0 else None

        values = [idx]
        for fi2, field in enumerate(FIELD_MAP[1:], 2):
            if field in ("ma_nv", "ho_va_ten", "phong_cap_1", "phong_cap_2",
                         "phong_cap_3", "vi_tri", "dia_diem"):
                values.append(emp.get(field, ""))
            else:
                values.append(s.get(field, ""))
        for k in extra_cols:
            values.append(s.get(k, extra_cols[k]))

        for ci, val in enumerate(values, 1):
            cell = ws1.cell(row=r, column=ci, value=val)
            cell.font   = _font(size=9)
            cell.border = bd
            if row_fill:
                cell.fill = row_fill
            if ci in TEXT_COLS:
                cell.alignment = la if ci in {3, 4, 5, 6, 7} else ca
            elif ci in NUM_INT_COLS:
                cell.alignment = ca
                cell.number_format = "0.0"
            elif ci in NUM_DEC_COLS:
                cell.alignment = ra
                cell.number_format = "0.00"
            else:
                cell.alignment = ca

        cong_val = s.get("cong_tinh_luong", 0)
        ws1.cell(row=r, column=22).fill = (
            _fill("C6EFCE") if cong_val >= STD_CONG else
            _fill("FFEB9C") if cong_val >= STD_CONG * 0.8 else
            _fill("FFC7CE")
        )

        for k in totals:
            totals[k] += s.get(k, 0)

    # Total row
    total_row = hdr_row + len(employees) + 1
    ws1.merge_cells(f"A{total_row}:H{total_row}")
    tc = ws1.cell(row=total_row, column=1, value="TỔNG CỘNG")
    tc.font = _font(bold=True); tc.fill = _fill(C_TOTAL)
    tc.alignment = ca; tc.border = bd

    total_values_map = {
        9:  ("",                           ""),
        10: (int(totals["nghi_phep_nguyen_ngay"]), "0"),
        11: (totals["nghi_phep_nua_ngay"], "0.0"),
        12: (int(totals["nghi_le"]),        "0"),
        13: (int(totals["wfh"]),            "0"),
        14: (int(totals["cong_tac"]),       "0"),
        15: (int(totals["nghi_khong_luong"]), "0"),
        16: (round(totals["tong_cong_thuc_te"], 2), "0.00"),
        17: (round(totals["tong_cong_huong_luong"], 2), "0.00"),
        18: (int(totals["dm_gt_4h"]),       "0"),
        19: (int(totals["dm_1h_4h"]),       "0"),
        20: (int(totals["phut_muon_lt_1h"]), "0"),
        21: (round(totals["tru_sm"], 4),    "0.0000"),
        22: (round(totals["cong_tinh_luong"], 2), "0.00"),
    }
    for ci, (val, fmt) in total_values_map.items():
        cell = ws1.cell(row=total_row, column=ci, value=val)
        cell.font = _font(bold=True); cell.fill = _fill(C_TOTAL)
        cell.alignment = ca; cell.border = bd
        if fmt:
            cell.number_format = fmt

    ws1.freeze_panes = f"I{hdr_row + 1}"

    # ═══════════════════════════════════════════
    # SHEET 2: CHI TIẾT NGÀY
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet("Chi tiết ngày")
    all_days_in_period = _get_all_days_in_period(period_start, period_end)
    n_days = len(all_days_in_period)
    total_detail_cols = 6 + n_days
    end_col2 = get_column_letter(total_detail_cols)

    ws2.merge_cells(f"A1:{end_col2}1")
    ws2["A1"] = f"CHI TIẾT CHẤM CÔNG THEO NGÀY — Kỳ: {ky_str}"
    ws2["A1"].font = _font(bold=True, color="1F4E79", size=13)
    ws2["A1"].alignment = ca

    FIXED_HDR2 = [
        ("STT",          5),
        ("Mã NV",        9),
        ("Họ và tên",    22),
        ("Phòng cấp 2",  18),
        ("Vị trí",       18),
        ("Ca làm việc",  14),
    ]
    for ci, (lbl, wid) in enumerate(FIXED_HDR2, 1):
        cell = ws2.cell(row=3, column=ci, value=lbl)
        cell.font = _font(bold=True, color="FFFFFF", size=9)
        cell.fill = _fill(C_HDR1); cell.alignment = ca; cell.border = bd
        ws2.column_dimensions[get_column_letter(ci)].width = wid

    for di, d in enumerate(all_days_in_period):
        col = 7 + di
        ds  = d.strftime("%Y-%m-%d")
        wday = WEEKDAY_VI[d.weekday()]
        is_we  = d.weekday() in off_wdays
        is_hol = ds in holidays
        fc = C_WE if is_we else (C_HOL if is_hol else C_HDR2)
        tc_color = "666666" if is_we else "FFFFFF"

        c3 = ws2.cell(row=3, column=col, value=wday)
        c3.font = _font(bold=True, color=tc_color, size=8)
        c3.fill = _fill(fc); c3.alignment = ca; c3.border = bd

        c4 = ws2.cell(row=4, column=col, value=f"{d.day:02d}/{d.month:02d}")
        c4.font = _font(bold=True, color=tc_color, size=8)
        c4.fill = _fill(fc); c4.alignment = ca; c4.border = bd
        ws2.column_dimensions[get_column_letter(col)].width = 9

    ws2.row_dimensions[3].height = 20
    ws2.row_dimensions[4].height = 20

    row_ptr = 5
    for idx, emp in enumerate(employees, 1):
        detail_map = {dd["date"]: dd for dd in emp.get("daily_detail", [])}
        work_start_emp, work_end_emp = _work_hours_for_location(emp.get("dia_diem", ""))
        ca_label = (
            f"{work_start_emp.strftime('%H:%M')} ~ {work_end_emp.strftime('%H:%M')}"
            + (" (HòaLạc)" if _is_hoa_lac(emp.get("dia_diem", "")) else "")
        )

        r_in  = row_ptr
        r_out = row_ptr + 1

        for ci in range(1, 7):
            ws2.merge_cells(start_row=r_in, start_column=ci, end_row=r_out, end_column=ci)

        ws2.cell(row=r_in, column=1, value=idx).alignment  = ca
        ws2.cell(row=r_in, column=2, value=emp["ma_nv"]).alignment = ca
        ws2.cell(row=r_in, column=3, value=emp["ho_va_ten"]).alignment = la
        ws2.cell(row=r_in, column=4, value=emp.get("phong_cap_2", emp["don_vi"])).alignment = la
        ws2.cell(row=r_in, column=5, value=emp["vi_tri"]).alignment  = la
        ws2.cell(row=r_in, column=6, value=ca_label).alignment       = ca

        for ci in range(1, 7):
            for r in (r_in, r_out):
                ws2.cell(row=r, column=ci).font   = _font(size=8)
                ws2.cell(row=r, column=ci).border = bd

        for di, d in enumerate(all_days_in_period):
            ds   = d.strftime("%Y-%m-%d")
            col  = 7 + di
            dd   = detail_map.get(ds, {})
            dtype = dd.get("day_type", "")
            ci_   = dd.get("check_in",  "-:-")
            co_   = dd.get("check_out", "-:-")
            note  = dd.get("note", "")
            leave_cong = dd.get("leave_cong", 0.0)

            is_we  = d.weekday() in off_wdays
            is_hol = ds in holidays

            # Màu nền
            if is_we:
                fc = C_WE
            elif is_hol:
                fc = C_HOL
            elif "Nghỉ phép nửa ngày" in dtype:
                fc = C_NP_H
            elif dtype == "Nghỉ phép/ốm":
                fc = C_NP
            elif dtype == "WFH":
                fc = C_WFH
            elif dtype == "Công tác":
                fc = C_CT
            elif dtype == "Nghỉ không lương":
                fc = C_NKL
            elif dtype in ("Đi làm", "Đi làm (cập nhật)"):
                fc = C_DL if not note else "FFF2CC"
            else:
                fc = "FFFFFF"

            # Hiển thị ô check-in
            if is_we:
                display_in = ""
            elif is_hol:
                display_in = "Lễ"
            elif dtype == "Nghỉ phép/ốm":
                display_in = "NP"
            elif "Nghỉ phép nửa ngày" in dtype:
                display_in = "NP½"
            elif dtype == "WFH":
                display_in = "WFH"
            elif dtype == "Công tác":
                display_in = "CT"
            elif dtype == "Nghỉ không lương":
                display_in = "NKL"
            else:
                display_in = ci_ if ci_ != "-:-" else ""

            # Hiển thị ô check-out
            if is_we or is_hol or dtype in ("Nghỉ phép/ốm", "WFH", "Công tác"):
                display_out = ""
            elif "Nghỉ phép nửa ngày" in dtype:
                display_out = co_ if co_ != "-:-" else ""
            elif dtype == "Nghỉ không lương":
                display_out = ""
            else:
                display_out = co_ if co_ != "-:-" else ""

            # Note hiển thị ở ô check-out nếu có
            if note and dtype not in ("Nghỉ tuần",):
                display_out = note[:12] if not display_out else display_out

            cell_in = ws2.cell(row=r_in,  column=col, value=display_in)
            cell_in.font = _font(size=8); cell_in.fill = _fill(fc)
            cell_in.alignment = ca; cell_in.border = bd

            cell_out = ws2.cell(row=r_out, column=col, value=display_out)
            cell_out.font = _font(size=8); cell_out.fill = _fill(fc)
            cell_out.alignment = ca; cell_out.border = bd

        ws2.row_dimensions[r_in].height  = 14
        ws2.row_dimensions[r_out].height = 14
        row_ptr += 2

    ws2.freeze_panes = "G5"

    # ═══════════════════════════════════════════
    # SHEET 3: CHÚ GIẢI
    # ═══════════════════════════════════════════
    ws3 = wb.create_sheet("Chú giải")
    ws3["A1"] = "CHÚ GIẢI MÀU SẮC VÀ KÝ HIỆU"
    ws3["A1"].font = _font(bold=True, size=11)

    legend = [
        (C_WE,   "Ngày nghỉ tuần (T7/CN)"),
        (C_HOL,  "Ngày nghỉ lễ"),
        (C_NP,   "Nghỉ phép/ốm cả ngày"),
        (C_NP_H, "Nghỉ phép nửa ngày (0.5 công)"),
        (C_WFH,  "Làm việc từ xa - WFH"),
        (C_CT,   "Đi công tác"),
        (C_NKL,  "Nghỉ không lương"),
        (C_DL,   "Đi làm bình thường"),
        ("FFF2CC", "Đi muộn/về sớm có phút tính trừ"),
        ("C6EFCE", "Công tính lương ≥ 26"),
        ("FFEB9C", "Công tính lương ≥ 20"),
        ("FFC7CE", "Công tính lương < 20"),
    ]
    for ri, (color, desc) in enumerate(legend, 3):
        ws3.cell(row=ri, column=1, value="  ").fill = _fill(color)
        ws3.cell(row=ri, column=2, value=desc).font = _font(size=9)

    ws3["A17"] = "CÔNG THỨC TÍNH"
    ws3["A17"].font = _font(bold=True)
    formulas = [
        ("Công hưởng lương",  "= 26 - Nghỉ KL + Nghỉ phép (kể cả 0.5) + Nghỉ lễ + WFH + Công tác"),
        ("Trừ SM",            "= (Ngày ĐM>4h × 1) + (Ngày ĐM 1-4h × 0.5) + (Tổng phút <1h / 480)"),
        ("Công tính lương",   "= Công hưởng lương - Trừ SM"),
        ("Nghỉ KL",           "Ngày không có CC hoặc làm < 2 tiếng"),
        ("Đi muộn/về sớm",   "Tối đa 2 đơn/kỳ, tổng ≤ 60 phút → được miễn trừ"),
        ("Ca Hòa Lạc",        "08:30 ~ 16:30 (nghỉ trưa 12:00 ~ 13:15)"),
        ("Ca thông thường",   "08:30 ~ 17:30 (nghỉ trưa 12:00 ~ 13:15)"),
        ("Nghỉ phép nửa ngày","so_ngay_nghi=0.5 hoặc thời gian < nửa ca → 0.5 công"),
        ("Cập nhật công",     "Đơn Cập nhật công đã duyệt → ghi đè giờ CC ngày đó"),
    ]
    for ri, (k, v) in enumerate(formulas, 18):
        ws3.cell(row=ri, column=1, value=k).font = _font(bold=True, size=9)
        ws3.cell(row=ri, column=2, value=v).font = _font(size=9)

    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 65

    # Save
    if not output_path:
        safe = re.sub(r"[^\w]", "_", filter_value or "all")
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"/tmp/bang_cham_cong_{year_month}_{now_str}_{safe}.xlsx"

    os.makedirs(
        os.path.dirname(output_path) if os.path.dirname(output_path) else ".",
        exist_ok=True,
    )
    wb.save(output_path)
    logger.info("Excel saved: %s (%d employees)", output_path, len(employees))

    return json.dumps({
        "status":       "success",
        "file_path":    output_path,
        "year_month":   year_month,
        "ky_cham_cong": ky_str,
        "so_nhan_vien": len(employees),
        "sheets":       ["Tổng hợp", "Chi tiết ngày", "Chú giải"],
        "message":      f"Đã xuất {len(employees)} NV, kỳ {ky_str}, tại: {output_path}",
    }, ensure_ascii=False)


@mcp.tool()
def compute_and_export(
    session_id:              str,
    year_month:              str,
    filter_type:             str = "all",
    filter_value:            str = "",
    company_code:            str = "HITC",
    output_path:             str = "",
    extra_columns:           str = "",
    custom_formula_notes:    str = "",
    data_overrides:          str = "",
) -> str:
    """All-in-one: tính toán bảng chấm công và xuất Excel ngay."""
    return export_attendance_excel(
        session_id=session_id, year_month=year_month,
        filter_type=filter_type, filter_value=filter_value,
        company_code=company_code, output_path=output_path,
        extra_columns=extra_columns,
        custom_formula_notes=custom_formula_notes,
        data_overrides=data_overrides,
    )


@mcp.tool()
def send_attendance_report(
    session_id:              str,
    year_month:              str,
    filter_type:             str       = "all",
    filter_value:            str       = "",
    to_emails:               list[str] = None,
    send_to_don_vi:          str       = "",
    subject:                 str       = "",
    body:                    str       = "",
    company_code:            str       = "HITC",
    extra_columns:           str       = "",
    custom_formula_notes:    str       = "",
    data_overrides:          str       = "",
) -> str:
    """Xuất bảng chấm công Excel và gửi email đính kèm."""
    if not _session_ok(session_id):
        return json.dumps({"error": "Session không hợp lệ."}, ensure_ascii=False)

    if not to_emails and not send_to_don_vi:
        return json.dumps({"error": "Cần truyền to_emails hoặc send_to_don_vi."}, ensure_ascii=False)

    export_res = json.loads(export_attendance_excel(
        session_id=session_id, year_month=year_month,
        filter_type=filter_type, filter_value=filter_value,
        company_code=company_code,
        extra_columns=extra_columns,
        custom_formula_notes=custom_formula_notes,
        data_overrides=data_overrides,
    ))
    if export_res.get("status") != "success":
        return json.dumps(export_res, ensure_ascii=False)

    file_path = export_res["file_path"]
    ky_str    = export_res["ky_cham_cong"]
    so_nv     = export_res["so_nhan_vien"]

    if not subject:
        subject = f"[{company_code}] Bảng chấm công tổng hợp — Kỳ {ky_str}"
    if not body:
        body = (
            f"Kính gửi,\n\nĐính kèm bảng chấm công tổng hợp kỳ {ky_str} ({so_nv} nhân viên).\n"
            f"File gồm 3 sheets: Tổng hợp, Chi tiết ngày, Chú giải.\n\nTrân trọng."
        )

    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        from email.utils import formataddr
        from app.core.config import settings

        db = get_db()
        recipients: list[str] = []
        for r in (to_emails or []):
            if "@" in r:
                recipients.append(r)
            else:
                nv = db["instance_data_thong_tin_nhan_vien"].find_one(
                    {"ten_dang_nhap": r, "is_deleted": {"$ne": True}}, {"email": 1}
                )
                if nv and nv.get("email"):
                    recipients.append(nv["email"])

        if send_to_don_vi:
            for nv in db["instance_data_thong_tin_nhan_vien"].find(
                {
                    "is_deleted": {"$ne": True}, "company_code": company_code,
                    "trang_thai_lao_dong.value": "Đang làm việc",
                    "$or": [
                        {"phong_cap_1.value":     send_to_don_vi},
                        {"phong_cap_3.value": send_to_don_vi},
                        {"phong_cap_2.value": send_to_don_vi},
                        {"path_phong_ban": {"$regex": send_to_don_vi, "$options": "i"}},
                    ],
                }, {"email": 1},
            ):
                if nv.get("email"):
                    recipients.append(nv["email"])

        recipients = list(set(recipients))
        if not recipients:
            return json.dumps({
                "status":    "exported_only",
                "file_path": file_path,
                "message":   "Xuất Excel OK nhưng không tìm thấy email để gửi.",
            }, ensure_ascii=False)

        msg            = MIMEMultipart()
        msg["Subject"] = subject
        msg["From"]    = formataddr(("MODATA AI System", settings.MAIL_FROM))
        msg["To"]      = ", ".join(recipients[:50])
        msg.attach(MIMEText(body, "plain", "utf-8"))

        with open(file_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",
                        f'attachment; filename="{os.path.basename(file_path)}"')
        msg.attach(part)

        smtp_cls = smtplib.SMTP_SSL if settings.MAIL_PORT == 465 else smtplib.SMTP
        with smtp_cls(settings.MAIL_HOST, settings.MAIL_PORT, timeout=30) as smtp:
            if settings.MAIL_PORT != 465 and settings.MAIL_USE_TLS:
                smtp.starttls()
            if settings.MAIL_USERNAME and settings.MAIL_PASSWORD:
                smtp.login(settings.MAIL_USERNAME, settings.MAIL_PASSWORD)
            smtp.sendmail(settings.MAIL_FROM, recipients, msg.as_string())

        return json.dumps({
            "status":       "sent",
            "file_path":    file_path,
            "recipients":   recipients,
            "so_nhan_vien": so_nv,
            "message":      f"Đã gửi bảng chấm công đến {len(recipients)} người.",
        }, ensure_ascii=False)

    except Exception as e:
        logger.error("send_attendance_report error: %s", e, exc_info=True)
        return json.dumps({
            "status":    "exported_only",
            "file_path": file_path,
            "message":   f"Xuất Excel OK nhưng gửi mail thất bại: {e}",
        }, ensure_ascii=False)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(mcp.http_app(), host="0.0.0.0", port=8020)